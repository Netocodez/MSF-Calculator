from flask import Flask, render_template, request, jsonify, send_file, send_from_directory, url_for, abort

import pandas as pd
import numpy as np
import os
import logging
from datetime import datetime
from dateutil import parser
import xlsxwriter
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from werkzeug.utils import safe_join
from openpyxl.styles import Font, Alignment, PatternFill

#from utilities import process_emr_data
import msf_common

# Flask app
app = Flask(__name__)

# Configuration
BASE_DIR = os.path.abspath(os.path.dirname(__file__))

UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
DOWNLOAD_FOLDER = os.path.join(BASE_DIR, "downloads")

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(DOWNLOAD_FOLDER, exist_ok=True)

app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["DOWNLOAD_FOLDER"] = DOWNLOAD_FOLDER

ALLOWED_EXTENSIONS = {'.xlsx', '.xls', '.csv'}

# Global variable for formatted reporting period
formatted_period = None

#columns to process
columns_to_read = [
    'State', 'LGA', 'FacilityName', 'PatientHospitalNo', 'PEPID', 'uuid', 'ARTStatus_PreviousQuarter','CurrentARTStatus', 'DOB', 'ARTStartDate', 'DateConfirmedHIV+', 'Pharmacy_LastPickupdate',
    'DateResultReceivedFacility', 'Date_Transfered_In','Whostage','CurrentCD4','Current_CD4_LFA_Result','Serology_for_CrAg_Result','CSF_for_CrAg_Result',
    'CurrentPregnancyStatus', 'First_TPT_Pickupdate', 'Current_TPT_Received', 'Current_TB_Status', 'CurrentRegimenLine',
    'DaysOfARVRefill', 'DSD_Model', 'Sex', 'KPType', 'Outcomes_Date', 'CurrentViralLoad', 'ViralLoadIndication', 'DateofCurrent_TBStatus'
]

b_columns_to_read = [
    'uuid', 'CurrentARTStatus'
]

r_columns_to_read = [
    'State', 'LGA', 'Facility', 'Hospital Number', 'Unique ID', 'Patient ID', 'Date of TPT Start (yyyy-mm-dd)', 'TPT Type', 'TPT Completion date (yyyy-mm-dd)'
]

# Columns
DATE_COLUMNS = [
    'DOB', 'ARTStartDate','DateConfirmedHIV+', 'Pharmacy_LastPickupdate',
    'DateResultReceivedFacility', 'Date_Transfered_In', 'Outcomes_Date', 'DateofCurrent_TBStatus', 'First_TPT_Pickupdate'
]

NUMERIC_COLUMNS = [
    'DaysOfARVRefill', 'CurrentViralLoad', 'CurrentCD4'
]

EMRfilename = "LAMISNMRS.csv"
emr_df = pd.read_csv(EMRfilename, encoding='utf-8')

print("emr_df columns:", emr_df.columns.tolist())

# Logging
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")


# Utility: check file extension
def is_allowed_file(filename):
    return os.path.splitext(filename)[1].lower() in ALLOWED_EXTENSIONS

# Utility: load file (CSV or Excel)
def load_file(file, columns_to_read=None):
    file_ext = os.path.splitext(file.filename)[1].lower()
    if file_ext == '.csv':
        return pd.read_csv(
            file,
            dtype=str,
            encoding='utf-8',
            lineterminator='\n',
            quotechar='"',
            escapechar='\\',
            usecols=columns_to_read,
            skip_blank_lines=True
        )
    elif file_ext in ['.xls', '.xlsx']:
        return pd.read_excel(file, sheet_name=0, dtype=object, usecols=columns_to_read, engine='openpyxl')
    else:
        raise ValueError("Unsupported file type")

# Utility: Clean and normalize patient/facility identifiers
def clean_id(val):
    if pd.isna(val):
        return ''
    val = str(val).strip().lower().replace(' ', '').lstrip('0')
    return val

def process_emr_data(df, dfbaseline, emr_df):
    # Remove rows with any blank fields in mapping
    emr_df = emr_df[(emr_df != '').all(axis=1)]
    
    # Select and deduplicate necessary columns from emr_df
    emr_subset = emr_df[['Name on NMRS', 'LGA', 'STATE', 'Name on Lamis']].drop_duplicates(subset='Name on NMRS')

    # Merge once using FacilityName <-> Name on NMRS
    df = df.merge(
        emr_subset,
        how='left',
        left_on='FacilityName',
        right_on='Name on NMRS',
        suffixes=('', '_emr')
    )

    # Fill missing LGA and State from EMR
    df['LGA'] = df['LGA'].fillna(df['LGA_emr'])
    df['State'] = df['State'].fillna(df['STATE'])

    # Replace FacilityName if different
    df.loc[df['Name on Lamis'] != df['FacilityName'], 'FacilityName'] = df['Name on Lamis']

    # Drop extra columns
    df.drop(['Name on NMRS', 'LGA_emr', 'STATE', 'Name on Lamis'], axis=1, inplace=True)

    # Normalize hospital numbers and unique IDs
    df['PatientHospitalNo1'] = df['PatientHospitalNo'].apply(clean_id)
    df['PatientUniqueID1'] = df['PEPID'].apply(clean_id)
    dfbaseline['Hospital Number1'] = dfbaseline['Hospital Number'].apply(clean_id)
    dfbaseline['Unique ID1'] = dfbaseline['Unique ID'].apply(clean_id)

    # Create consistent unique identifiers for both datasets
    dfbaseline['unique identifiers'] = (
        dfbaseline["LGA"].astype(str).str.lower().str.strip().str.replace(' ', '') +
        dfbaseline["Facility"].astype(str).str.lower().str.strip().str.replace(' ', '') +
        dfbaseline["Hospital Number1"] +
        dfbaseline["Unique ID1"]
    )

    df['unique identifiers'] = (
        df["LGA"].astype(str).str.lower().str.strip().str.replace(' ', '') +
        df["FacilityName"].astype(str).str.lower().str.strip().str.replace(' ', '') +
        df["PatientHospitalNo1"] +
        df["PatientUniqueID1"]
    )

    # Drop duplicates from baseline data
    dfbaseline = dfbaseline.drop_duplicates(subset=['unique identifiers'], keep=False)

    # Identify duplicates in 'unique identifiers'
    dup_mask = df.duplicated('unique identifiers', keep=False)

    # Only modify duplicates
    df.loc[dup_mask, 'unique identifiers'] = (
        df.loc[dup_mask]
        .groupby('unique identifiers')
        .cumcount()
        .astype(str)
        .radd(df.loc[dup_mask, 'unique identifiers'] + '_')
    )

    # Merge into df
    df = df.merge(
        dfbaseline[['unique identifiers', 'Date of TPT Start (yyyy-mm-dd)', 'TPT Type']],
        on='unique identifiers',
        how='left',
        suffixes=('', '_baseline')
    )
    #df.to_excel('df.xlsx')

    # Fill missing TPT values
    df['Date of TPT Start (yyyy-mm-dd)'] = pd.to_datetime(df['Date of TPT Start (yyyy-mm-dd)'], errors='coerce', dayfirst=True)
    df['First_TPT_Pickupdate'] = pd.to_datetime(df['First_TPT_Pickupdate'], errors='coerce', dayfirst=True)
    df['First_TPT_Pickupdate'] = df['First_TPT_Pickupdate'].fillna(df['Date of TPT Start (yyyy-mm-dd)'])
    df['Current_TPT_Received'] = df['Current_TPT_Received'].fillna(df['TPT Type'])

    return df

@app.route('/')
def home():
    return render_template('index.html')

@app.route('/fetch', methods=['POST'])
def fetch_data():
    
    OLD_MSF_FOLDER = os.path.join(app.config["UPLOAD_FOLDER"], "old_msf")
    os.makedirs(OLD_MSF_FOLDER, exist_ok=True)
    
    file1 = request.files.get("file1")
    file2 = request.files.get("file2")
    file3 = request.files.get("file3")

    if not file1 or not is_allowed_file(file1.filename):
        return jsonify({"message": "Current ART Line List must be a CSV or Excel file."}), 400

    if file2 and not is_allowed_file(file2.filename):
        return jsonify({"message": "Baseline ART Line List must be a CSV or Excel file."}), 400
    
    if file3 and not is_allowed_file(file3.filename):
        return jsonify({"message": "Baseline Lamis Radet must be a CSV or Excel file."}), 400

    try:
        try:
            # Load and clean current ART line list
            df = load_file(file1, columns_to_read=columns_to_read)
            #df = clean_dataframe(df)

            # Merge baseline ART data if provided
            if file2:
                df_baseline = load_file(file2, columns_to_read=b_columns_to_read)
                if 'uuid' in df.columns and 'uuid' in df_baseline.columns and 'CurrentARTStatus' in df_baseline.columns:
                    df = df.merge(
                        df_baseline[['uuid', 'CurrentARTStatus']],
                        on='uuid', how='left', suffixes=('', '_baseline')
                    )
                    df['ARTStatus_PreviousQuarter'] = df['CurrentARTStatus_baseline']
                    
            if file3:
                #emr_df = pd.read_excel(EMRfilename, sheet_name=0)
                #emr_df = load_file(EMRfilename, columns_to_read=None)
                dfbaselineRadet = load_file(file3, columns_to_read=r_columns_to_read)
                df = process_emr_data(df, dfbaselineRadet, emr_df)
                    
            #df.to_excel('df.xlsx')
            for col in DATE_COLUMNS:
                if col in df.columns:
                    df[col] = pd.to_datetime(df[col], errors='coerce', dayfirst=True)
                    
            for col in NUMERIC_COLUMNS:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce')

            # Read start and end dates from form data
            end_date = request.form.get("endDate")

            global formatted_period
            if end_date:
                end_date = pd.to_datetime(end_date)
                formatted_period = end_date.to_period('M').strftime('%B %Y')
                Period = end_date.to_period('M')  # Add this line

            # your code
            #data processing logic here...
            bins = [-np.inf, 1, 4, 9, 14, 19, 24, 29, 34, 39, 44, 49, 54, 59, 64, np.inf]
            labels = ['<1', '1-4', '5-9', '10-14', '15-19', '20-24', '25-29', '30-34', 
                    '35-39', '40-44', '45-49', '50-54', '55-59', '60-64', '65+']

            bins2 = [0, 14, float('inf')]
            labels2 = ['<15', '>=15']
            
            AGE_BANDS = [
                '<1', '1-4', '5-9', '10-14',
                '15-19', '20-24', '25-29',
                '30-34', '35-39', '40-44',
                '45-49', '50-54', '55-59',
                '60-64', '65+'
            ]

            def standardize_ageband_pivot(pt):
                pt = pt.reindex(columns=AGE_BANDS, fill_value=0)
                pt = pt.reindex(['M', 'F'], fill_value=0)
                pt.rename(index={'M': 'Male', 'F': 'Female'}, inplace=True)
                return pt

            # Extract unique facility names as a list
            unique_facilities = df['FacilityName'].unique()
            facilities_text = ', '.join(unique_facilities)
            print(facilities_text)
            
            #function to calculate current age and mirror excel datedif function
            def calculate_age_vectorized(df, dob_col='DOB', ref_date=None):
                # ✅ pick the reference date
                if ref_date is None:
                    today = pd.Timestamp.today().normalize()  # current day
                else:
                    today = pd.to_datetime(ref_date) + pd.offsets.MonthEnd(0)  # last day of month for given date

                # ✅ fully vectorized age calculation
                dob = df[dob_col]
                age = (today.year - dob.dt.year 
                    - ((dob.dt.month > today.month) | 
                        ((dob.dt.month == today.month) & (dob.dt.day > today.day))).astype(int))

                return age
            
            df['Age'] = calculate_age_vectorized(df, 'DOB', ref_date=end_date)

            df['Age Band'] = pd.cut(df['Age'], bins=bins, labels=labels, right=True)
            df['Age Band2'] = pd.cut(df['Age'], bins=bins2, labels=labels2)
            last_year = (Period.to_timestamp() - pd.DateOffset(months=12)).to_period('M')
            last_6mths = (Period.to_timestamp() - pd.DateOffset(months=6)).to_period('M')
            #df.to_excel('df.xlsx')

            # Creating additional columns for pregnant, breastfeeding, and TB clients
            df['IsPregnant'] = df['CurrentPregnancyStatus'].apply(lambda x: 1 if x == "Pregnant" else 0)
            df['IsBreastfeeding'] = df['CurrentPregnancyStatus'].apply(lambda x: 1 if x == "Breastfeeding" else 0)
            df['OnTB'] = df['Current_TB_Status'].apply(lambda x: 1 if x == "On treatment for disease" else 0)

            # Creating columns for regimen line
            df['IsFirstLine'] = df['CurrentRegimenLine'].apply(lambda x: 1 if x in ['Adult 1st line ARV regimen', 'Child 1st line ARV regimen'] else 0)
            df['IsSecondLine'] = df['CurrentRegimenLine'].apply(lambda x: 1 if x in ['Adult 2nd line ARV regimen', 'Child 2nd line ARV regimen'] else 0)
            df['isThirdLine'] = df['CurrentRegimenLine'].apply(lambda x: 1 if x  in ['Adult 3rd Line ARV Regimens', 'Child 3rd line ARV regimen'] else 0)

            # Creating MMD columns
            df['IsMMD3'] = df['DaysOfARVRefill'].apply(lambda x: 1 if x == 90 else 0)
            df['IsMMD4'] = df['DaysOfARVRefill'].apply(lambda x: 1 if x == 120 else 0)
            df['IsMMD5'] = df['DaysOfARVRefill'].apply(lambda x: 1 if x == 150 else 0)
            df['IsMMD6'] = df['DaysOfARVRefill'].apply(lambda x: 1 if x == 180 else 0)

            # Creating DSD Models columns
            df['IsFacilityModel'] = df['DSD_Model'].apply(lambda x: 1 if x == 'Facility Dispensing' else 0)
            df['IsCommunityModel'] = df['DSD_Model'].apply(lambda x: 1 if x == 'Decentralized Drug Delivery (DDD)' else 0)
        
        except Exception as e:
            logging.exception("Error Preparing file for Analysis")
            return jsonify({'error': str(e)}), 500
        
        
        try:
            #ART 2 (Newly Started on ART
            # Filter only active clients
            df_TxNew = df[df['ARTStartDate'].dt.to_period('M') == Period].copy()

            # Use .loc to safely set values in the subset
            df_TxNew.loc[:, 'TxNew'] = df_TxNew['CurrentARTStatus'].apply(lambda x: 1 if x == "Active" else 0)

            # Creating the original summary pivot table
            ART2Summary = df_TxNew.pivot_table(
                index='Sex',
                columns='Age Band',
                values='TxNew',
                aggfunc='sum',
                fill_value=0,
                observed=False
            )

            # Reindexing and renaming Sex categories
            #ART2Summary = ART2Summary.reindex(['M', 'F'])
            #ART2Summary.rename(index={'M': 'Male', 'F': 'Female'}, inplace=True)
            
            ART2Summary = standardize_ageband_pivot(ART2Summary)

            # Adding 'Total' column
            ART2Summary['Total'] = ART2Summary.sum(axis=1)

            # Summing new clients and the new categories
            sum_pregnant = df_TxNew[df_TxNew['IsPregnant'] == 1]['TxNew'].sum()
            sum_breastfeeding = df_TxNew[df_TxNew['IsBreastfeeding'] == 1]['TxNew'].sum()
            sum_on_tb = df_TxNew[df_TxNew['OnTB'] == 1]['TxNew'].sum()

            # Adding new rows for Pregnant, Breastfeeding, and TB Patients with NaN (blank) values
            ART2Summary.loc['Pregnant', :] = np.nan
            ART2Summary.loc['Breastfeeding', :] = np.nan
            ART2Summary.loc['TB Patients', :] = np.nan

            # Adding total sums to the new rows
            ART2Summary.loc['Pregnant', 'Total'] = sum_pregnant
            ART2Summary.loc['Breastfeeding', 'Total'] = sum_breastfeeding
            ART2Summary.loc['TB Patients', 'Total'] = sum_on_tb

            # Display the modified summary
            #ART2Summary
            
        except Exception as e:
            logging.exception("Error Processing ART 2 Summary")
            return jsonify({'error': str(e)}), 500
        
        
        try:
            #ART 3 (Current on ART)
            # Filter only active clients
            df_active = df[df['CurrentARTStatus'] == "Active"].copy()
            df_active.loc[:, 'TCS1'] = df_active['CurrentARTStatus'].apply(lambda x: 1 if x == "Active" else 0)

            # Creating the original ART3Summary pivot table
            ART3Summary = df_active.pivot_table(
                index='Sex',
                columns='Age Band',
                values='TCS1',
                aggfunc='sum',
                fill_value=0,
                observed=False
            )

            #ART3Summary = ART3Summary.reindex(['M', 'F'])
            #ART3Summary.rename(index={'M': 'Male', 'F': 'Female'}, inplace=True)
            ART3Summary = standardize_ageband_pivot(ART3Summary)
            ART3Summary['Total'] = ART3Summary.sum(axis=1)

            # Summing TCS1 for only active clients and the new categories
            sum_pregnant = df_active[df_active['IsPregnant'] == 1]['TCS1'].sum()
            sum_breastfeeding = df_active[df_active['IsBreastfeeding'] == 1]['TCS1'].sum()

            # Adding new rows with NaN (blank) values to avoid dtype issues
            ART3Summary.loc['Pregnant (subset of ART 3)', :] = np.nan
            ART3Summary.loc['Breastfeeding (subset of ART 3)', :] = np.nan

            # Adding new rows for Pregnant, Breastfeeding, and OnTB as total sums
            ART3Summary.loc['Pregnant (subset of ART 3)', 'Total'] = sum_pregnant
            ART3Summary.loc['Breastfeeding (subset of ART 3)', 'Total'] = sum_breastfeeding

            # Display the modified ART3Summary
            #ART3Summary
            
            #ART3 CONTD Regimen Lines
            def get_group_summary(df_sub, label):
                pt = df_sub.pivot_table(
                    index=None,
                    columns=['Sex', 'Age Band2'],
                    values='TCS1',
                    aggfunc='sum',
                    fill_value=0,
                    observed=False
                )
                pt = pt.reindex(columns=pd.MultiIndex.from_product([['M', 'F'], ['<15', '>=15']]), fill_value=0)
                pt.columns.names = ['Sex', 'Age Group']
                pt.rename(columns={'M': 'Male', 'F': 'Female'}, level=0, inplace=True)

                # Create a one-row DataFrame labeled with 'label'
                row = pt.iloc[0] if not pt.empty else [0] * pt.shape[1]
                return pd.DataFrame([row], index=[label])

            regimenlines = []

            # Call function and append results
            regimenlines.append(get_group_summary(df_active[df_active['IsFirstLine'] == 1], '1st Line (subset of ART 3)'))
            regimenlines.append(get_group_summary(df_active[df_active['IsSecondLine'] == 1], '2nd Line (subset of ART 3)'))
            regimenlines.append(get_group_summary(df_active[df_active['isThirdLine'] == 1], '3rd Line (subset of ART 3)'))

            # Combine all into one DataFrame
            ART3aSummary = pd.concat(regimenlines)
            total_row = ART3aSummary.sum(numeric_only=True).to_frame().T
            total_row.index = ['Total']
            ART3aSummary = pd.concat([ART3aSummary, total_row])
            #ART3aSummary
            
            
            #ART 3 CONTD MMDs
            MMDs = []
            MMDs.append(get_group_summary(df_active[df_active['IsMMD3'] == 1], 'MMD3'))
            MMDs.append(get_group_summary(df_active[df_active['IsMMD4'] == 1], 'MMD4'))
            MMDs.append(get_group_summary(df_active[df_active['IsMMD5'] == 1], 'MMD5'))
            MMDs.append(get_group_summary(df_active[df_active['IsMMD6'] == 1], 'MMD6'))

            ART3bSummary = pd.concat(MMDs)
            total_row = ART3bSummary.sum(numeric_only=True).to_frame().T
            total_row.index = ['Total']
            ART3bSummary = pd.concat([ART3bSummary, total_row])
            #ART3bSummary
            
            
            #ART3 CONTD MODELS
            ART3cSummary = df_active.pivot_table(
                index='DSD_Model',
                columns=None,
                values='TCS1',
                aggfunc='sum',
                fill_value=0,
                observed=False
            )

            ART3cSummary.rename(index={'Facility Dispensing': 'Facility-based models', 'Decentralized Drug Delivery (DDD)': 'Community-based models'}, inplace=True)
            ART3cSummary = ART3cSummary.reindex(['Facility-based models', 'Community-based models'])

            #ART3cSummary
        except Exception as e:
            logging.exception("Error Processing ART 3 Summary")
            return jsonify({'error': str(e)}), 500
        
        
        try:
            #ART 5 (Current on ART)
            # Ensure the 'Pharmacy_LastPickupdate' column is in datetime format and fill NaNs with a specific date
            df['Pharmacy_LastPickupdate2'] = pd.to_datetime(df['Pharmacy_LastPickupdate'], errors='coerce').fillna(pd.to_datetime('1900'))

            #Fill zero if the column contains number greater than 180
            df['DaysOfARVRefill2'] = df['DaysOfARVRefill'].apply(lambda x: 0 if x > 180 else x)

            # Calculate the 'NextAppt' column by adding the 'DaysOfARVRefill2' to 'Pharmacy_LastPickupdate2'
            df['NextAppt'] = df['Pharmacy_LastPickupdate2'] + pd.to_timedelta(df['DaysOfARVRefill2'], unit='D') 
            df['IITDate2'] = (df['NextAppt'] + pd.Timedelta(days=29)).fillna('1900')

            df['Losses date'] = df['Outcomes_Date']
            df.loc[df['Losses date'].isna(), 'Losses date'] = df['IITDate2']
            df['Losses date'] = pd.to_datetime(df['Losses date'])

            df['Loss Month'] = df['Losses date'].dt.to_period('M')

            # Filter for losses within the month
            df_Losses = df[
                df['CurrentARTStatus'].isin(["Death", "Transferred out", "LTFU", "Discontinued Care"]) &
                (df['Losses date'].dt.to_period('M') == Period)
            ].copy()

            df_Losses.loc[:, 'Losses'] = df_Losses['CurrentARTStatus'].apply(
                lambda x: 1 if x in ["Death", "Transferred out", "LTFU", "Discontinued Care"] else 0
            )

            df_Losses.loc[:, 'Dead'] = df_Losses['CurrentARTStatus'].apply(lambda x: 1 if x == "Death" else 0)
            df_Losses.loc[:, 'Transferred out'] = df_Losses['CurrentARTStatus'].apply(lambda x: 1 if x == "Transferred out" else 0)
            df_Losses.loc[:, 'LTFU'] = df_Losses['CurrentARTStatus'].apply(lambda x: 1 if x == "LTFU" else 0)
            df_Losses.loc[:, 'Discontinued Care'] = df_Losses['CurrentARTStatus'].apply(lambda x: 1 if x == "Discontinued Care" else 0)

            # Creating the original summary pivot table
            ART5summary = df_Losses.pivot_table(
                index='Sex',
                columns='Age Band',
                values='Losses',
                aggfunc='sum',
                fill_value=0,
                observed=False
            )
            
            ART5summary = standardize_ageband_pivot(ART5summary)
            ART5summary['Total'] = ART5summary.sum(axis=1)

            #ART5summary = ART5summary.reindex(['M', 'F'])
            #ART5summary.rename(index={'M': 'Male', 'F': 'Female'}, inplace=True)
            #ART5summary['Total'] = ART5summary.sum(axis=1)

            # Summing TCS1 for only active clients and the new categories
            sum_TO = df_Losses[df_Losses['Transferred out'] == 1]['Losses'].sum()
            sum_Stopped = df_Losses[df_Losses['Discontinued Care'] == 1]['Losses'].sum()
            sum_LTFU = df_Losses[df_Losses['LTFU'] == 1]['Losses'].sum()
            sum_Dead = df_Losses[df_Losses['Dead'] == 1]['Losses'].sum()

            # Adding new rows with NaN (blank) values to avoid dtype issues
            ART5summary.loc['Transferred out (subset of ART 5)', :] = np.nan
            ART5summary.loc['Stopped treatment (subset of ART 5)', :] = np.nan
            ART5summary.loc['Lost to follow up (subset of ART 5)', :] = np.nan
            ART5summary.loc['Dead (subset of ART 5)', :] = np.nan

            # Adding new rows for Pregnant, Breastfeeding, and OnTB as total sums
            ART5summary.loc['Transferred out (subset of ART 5)', 'Total'] = sum_TO
            ART5summary.loc['Stopped treatment (subset of ART 5)', 'Total'] = sum_Stopped
            ART5summary.loc['Lost to follow up (subset of ART 5)', 'Total'] = sum_LTFU
            ART5summary.loc['Dead (subset of ART 5)', 'Total'] = sum_Dead

            # Display the modified summary
            #ART5summary
            
        except Exception as e:
            logging.exception("Error Processing ART 5 Summary")
            return jsonify({'error': str(e)}), 500
        
        
        try: 
            #ART 6 (VL Routine)
            # Filter only active clients
            df['ARTStartDate'] = pd.to_datetime(df['ARTStartDate'])

            df_VL = df[
                (df['CurrentARTStatus'] == "Active") &
                (df['DateResultReceivedFacility'].dt.to_period('M') > last_year) &
                (df['ARTStartDate'].dt.to_period('M') <= last_6mths)
            ].copy()

            #df_VL.loc[:, 'VLRoutine'] = df_VL['ViralLoadIndication'].apply(lambda x: 1 if (x in ['Normal priority (status)', 'Initial', 'PMTCT, 32 - 36 weeks gestation']) or pd.isna(x) else 0)
            df_VL['VLRoutine'] = df_VL['ViralLoadIndication'].apply(
                lambda x: 1 if pd.isna(x) or str(x) in ['Normal priority (status)', 'Initial', 'PMTCT, 32 - 36 weeks gestation'] else 0
            )
            df_VL.loc[:, 'VLTargeted'] = df_VL['ViralLoadIndication'].apply(lambda x: 1 if x in ['Repeat', 'Confirmation', 'Immunologic failure', 'Clinical failure'] else 0)


            # Creating the original summary pivot table
            VLRoutine = df_VL.pivot_table(
                index='Sex',
                columns='Age Band',
                values='VLRoutine',
                aggfunc='sum',
                fill_value=0,
                observed=False
            )

            #VLRoutine = VLRoutine.reindex(['M', 'F'])
            #VLRoutine.rename(index={'M': 'Male', 'F': 'Female'}, inplace=True)
            VLRoutine = standardize_ageband_pivot(VLRoutine)
            VLRoutine['Total'] = VLRoutine.sum(axis=1)

            # Summing new categories
            sum_pregnant = df_VL[df_VL['IsPregnant'] == 1]['VLRoutine'].sum()
            sum_breastfeeding = df_VL[df_VL['IsBreastfeeding'] == 1]['VLRoutine'].sum()

            # Adding new rows with NaN (blank) values to avoid dtype issues
            VLRoutine.loc['Pregnant (subset of ART 6)', :] = np.nan
            VLRoutine.loc['Breastfeeding (subset of ART 6)', :] = np.nan

            # Adding new rows for Pregnant, Breastfeeding, and OnTB as total sums
            VLRoutine.loc['Pregnant (subset of ART 6)', 'Total'] = sum_pregnant
            VLRoutine.loc['Breastfeeding (subset of ART 6)', 'Total'] = sum_breastfeeding

            # Display the modified summary
            #VLRoutine
            
            
            #ART 6 VL Targeted
            # Creating the original summary pivot table
            VLTargeted = df_VL.pivot_table(
                index='Sex',
                columns='Age Band',
                values='VLTargeted',
                aggfunc='sum',
                fill_value=0,
                observed=False
            )

            #VLTargeted = VLTargeted.reindex(['M', 'F'])
            #VLTargeted.rename(index={'M': 'Male', 'F': 'Female'}, inplace=True)
            VLTargeted = standardize_ageband_pivot(VLTargeted)
            VLTargeted['Total'] = VLTargeted.sum(axis=1)

            # Summing new categories
            sum_pregnant = df_VL[df_VL['IsPregnant'] == 1]['VLTargeted'].sum()
            sum_breastfeeding = df_VL[df_VL['IsBreastfeeding'] == 1]['VLTargeted'].sum()
            sum_Fac_model = df_VL[df_VL['IsFacilityModel'] == 1]['VLTargeted'].sum()
            sum_cot_model = df_VL[df_VL['IsCommunityModel'] == 1]['VLTargeted'].sum()

            # Adding new rows for Pregnant, Breastfeeding, and OnTB as total sums
            VLTargeted.loc['Pregnant (subset of ART 6)', 'Total'] = sum_pregnant
            VLTargeted.loc['Breastfeeding (subset of ART 6)', 'Total'] = sum_breastfeeding
            VLTargeted.loc['Facility-based DSD models (subset of ART 6)', 'Total'] = sum_pregnant
            VLTargeted.loc['Community-based DSD models (subset of ART 6)', 'Total'] = sum_breastfeeding

            # Display the modified summary
            #VLTargeted
            
            
            #ART 7 (VL Routine Suppressed)
            # Filter only active clients
            df_VL_Sup = df[
                (df['CurrentARTStatus'] == "Active") &
                (df['DateResultReceivedFacility'].dt.to_period('M') > last_year) &
                (df['ARTStartDate'].dt.to_period('M') <= last_6mths) &
                (df['CurrentViralLoad'] < 1000)
            ].copy()

            #df_VL_Sup.loc[:, 'VLRoutine_Sup'] = df_VL_Sup['ViralLoadIndication'].apply(lambda x: 1 if (x in ['Normal priority (status)', 'Initial', 'PMTCT, 32 - 36 weeks gestation']) or pd.isna(x) else 0)
            df_VL_Sup.loc[:, 'VLRoutine_Sup'] = df_VL_Sup['ViralLoadIndication'].apply(
                lambda x: 1 if pd.isna(x) or str(x) in ['Normal priority (status)', 'Initial', 'PMTCT, 32 - 36 weeks gestation'] else 0
            )
            df_VL_Sup.loc[:, 'VLTargeted_Sup'] = df_VL_Sup['ViralLoadIndication'].apply(lambda x: 1 if x in ['Repeat', 'Confirmation', 'Immunologic failure', 'Clinical failure'] else 0)


            # Creating the original summary pivot table
            VLRoutine_Sup = df_VL_Sup.pivot_table(
                index='Sex',
                columns='Age Band',
                values='VLRoutine_Sup',
                aggfunc='sum',
                fill_value=0,
                observed=False
            )

            VLRoutine_Sup = VLRoutine_Sup.reindex(['M', 'F'])
            VLRoutine_Sup.rename(index={'M': 'Male', 'F': 'Female'}, inplace=True)
            VLRoutine_Sup['Total'] = VLRoutine_Sup.sum(axis=1)

            # Summing new categories
            sum_pregnant = df_VL_Sup[df_VL_Sup['IsPregnant'] == 1]['VLRoutine_Sup'].sum()
            sum_breastfeeding = df_VL_Sup[df_VL_Sup['IsBreastfeeding'] == 1]['VLRoutine_Sup'].sum()

            # Adding new rows with NaN (blank) values to avoid dtype issues
            VLRoutine_Sup.loc['Pregnant (subset of ART 6)', :] = np.nan
            VLRoutine_Sup.loc['Breastfeeding (subset of ART 6)', :] = np.nan

            # Adding new rows for Pregnant, Breastfeeding, and OnTB as total sums
            VLRoutine_Sup.loc['Pregnant (subset of ART 6)', 'Total'] = sum_pregnant
            VLRoutine_Sup.loc['Breastfeeding (subset of ART 6)', 'Total'] = sum_breastfeeding

            # Display the modified summary
            #VLRoutine_Sup
            
            
            #ART 7 VL Targeted Suppressed
            # Creating the original summary pivot table
            VLTargeted_Sup = df_VL_Sup.pivot_table(
                index='Sex',
                columns='Age Band',
                values='VLTargeted_Sup',
                aggfunc='sum',
                fill_value=0,
                observed=False
            )

            #VLTargeted_Sup = VLTargeted_Sup.reindex(['M', 'F'])
            #VLTargeted_Sup.rename(index={'M': 'Male', 'F': 'Female'}, inplace=True)
            VLTargeted_Sup = standardize_ageband_pivot(VLTargeted_Sup)
            VLTargeted_Sup['Total'] = VLTargeted_Sup.sum(axis=1)

            # Summing new categories
            sum_pregnant = df_VL_Sup[df_VL_Sup['IsPregnant'] == 1]['VLTargeted_Sup'].sum()
            sum_breastfeeding = df_VL_Sup[df_VL_Sup['IsBreastfeeding'] == 1]['VLTargeted_Sup'].sum()
            sum_Fac_model = df_VL_Sup[df_VL_Sup['IsFacilityModel'] == 1]['VLTargeted_Sup'].sum()
            sum_cot_model = df_VL_Sup[df_VL_Sup['IsCommunityModel'] == 1]['VLTargeted_Sup'].sum()

            # Adding new rows for Pregnant, Breastfeeding, and OnTB as total sums
            VLTargeted_Sup.loc['Pregnant (subset of ART 6)', 'Total'] = sum_pregnant
            VLTargeted_Sup.loc['Breastfeeding (subset of ART 6)', 'Total'] = sum_breastfeeding
            VLTargeted_Sup.loc['Facility-based DSD models (subset of ART 6)', 'Total'] = sum_pregnant
            VLTargeted_Sup.loc['Community-based DSD models (subset of ART 6)', 'Total'] = sum_breastfeeding

            # Display the modified summary
            #VLTargeted_Sup
            
            
            #ART 8 (Restart)
            df['ARTStartDate'] = pd.to_datetime(df['ARTStartDate'])

            # Filter only active clients
            #df_Restart = df[(df['CurrentARTStatus'] == "Active") & 
                        #((df['ARTStatus_PreviousQuarter'] != "Active") & (df['ARTStatus_PreviousQuarter'].notna())) &
                        #(df['ARTStartDate'].dt.to_period('M') != Period)].copy()
            #df_Restart.loc[:, 'Restart'] = df_Restart['Date_Transfered_In'].dt.to_period('M').apply(lambda x: 1 if x != Period else 0)
            
            # Step 1: Filter only active clients (and valid ARTStartDate)
            df_Restart = df[
                (df['CurrentARTStatus'] == "Active") &
                (df['ARTStartDate'].dt.to_period('M') != Period)
            ].copy()

            # Step 2: Define Restart condition based on transfer date and ART status in previous quarter
            df_Restart['Restart'] = np.where(
                (df_Restart['ARTStatus_PreviousQuarter'].notna()) &
                (df_Restart['ARTStatus_PreviousQuarter'] != "Active") &
                (df_Restart['Date_Transfered_In'].dt.to_period('M') != Period),
                1,
                0
            )         
            

            # Creating the original ART3Summary pivot table
            ART8Summary = df_Restart.pivot_table(
                index='Sex',
                columns='Age Band',
                values='Restart',
                aggfunc='sum',
                fill_value=0,
                observed=False
            )

            #ART8Summary = ART8Summary.reindex(['M', 'F'])
            #ART8Summary.rename(index={'M': 'Male', 'F': 'Female'}, inplace=True)
            ART8Summary = standardize_ageband_pivot(ART8Summary)
            ART8Summary['Total'] = ART8Summary.sum(axis=1)

            # Display the modified ART8Summary
            #ART8Summary
        
        except Exception as e:
            logging.exception("Error Processing ART 6 and 7 Summary")
            return jsonify({'error': str(e)}), 500
        
        
        try:
            #ART 9 (Transfer In)
            # Filter only active clients
            df_TI = df[(df['CurrentARTStatus'] == "Active")].copy()

            #df_TI.loc[:, 'TI'] = df_TI['CurrentARTStatus'].apply(lambda x: 1 if x == "Active" else 0)
            df_TI.loc[:, 'TI'] = df_TI['Date_Transfered_In'].dt.to_period('M').apply(lambda x: 1 if x == Period else 0)

            # Creating the original ART3Summary pivot table
            ART9Summary = df_TI.pivot_table(
                index='Sex',
                columns='Age Band',
                values='TI',
                aggfunc='sum',
                fill_value=0,
                observed=False
            )

            ART9Summary = ART9Summary.reindex(['M', 'F'])
            ART9Summary.rename(index={'M': 'Male', 'F': 'Female'}, inplace=True)
            ART9Summary['Total'] = ART9Summary.sum(axis=1)

            # Display the modified ART8Summary
            #ART9Summary
        
        except Exception as e:
            logging.exception("Error Processing ART 9 Summary")
            return jsonify({'error': str(e)}), 500
        
        
        try:
            #ART 10 (TB Screening Newly Initiated on ART)
            # Filter only active clients
            df_TBScrn = df[(df['CurrentARTStatus'] == "Active") & 
                        (df['DateofCurrent_TBStatus'].notna())].copy()
            df_TBScrn.loc[:, 'TBScrnNew'] = df_TBScrn['ARTStartDate'].dt.to_period('M').apply(lambda x: 1 if x == Period else 0)

            # Creating the original ART3Summary pivot table
            ART10aSummary = df_TBScrn.pivot_table(
                index='Sex',
                columns='Age Band',
                values='TBScrnNew',
                aggfunc='sum',
                fill_value=0,
                observed=False
            )

            ART10aSummary = ART10aSummary.reindex(['M', 'F'])
            ART10aSummary.rename(index={'M': 'Male', 'F': 'Female'}, inplace=True)
            ART10aSummary['Total'] = ART10aSummary.sum(axis=1)

            # Display the modified ART8Summary
            #ART10aSummary
            
            
            #ART 10b (TB Screening Previously on ART)
            # Filter only active clients
            df_TBScrnPrev = df[(df['CurrentARTStatus'] == "Active") & 
                        #(df['DateofCurrent_TBStatus'].notna()) &
                        (df['Pharmacy_LastPickupdate'].dt.to_period('M') == Period)].copy()
            df_TBScrnPrev.loc[:, 'TBScrnPrev'] = df_TBScrnPrev['ARTStartDate'].dt.to_period('M').apply(lambda x: 1 if x != Period else 0)

            # Creating the original ART3Summary pivot table
            ART10bSummary = df_TBScrnPrev.pivot_table(
                index='Sex',
                columns='Age Band',
                values='TBScrnPrev',
                aggfunc='sum',
                fill_value=0,
                observed=False
            )

            ART10bSummary = ART10bSummary.reindex(['M', 'F'])
            ART10bSummary.rename(index={'M': 'Male', 'F': 'Female'}, inplace=True)
            ART10bSummary['Total'] = ART10bSummary.sum(axis=1)

            # Display the modified ART8Summary
            #ART10bSummary
            
        except Exception as e:
            logging.exception("Error Processing ART 10 Summary")
            return jsonify({'error': str(e)}), 500
        
        try:
            df_ART19 = df[
                df['DateConfirmedHIV+'].dt.to_period('M') == Period
            ].copy()

            df_ART19['CurrentCD4'] = pd.to_numeric(
                df_ART19['CurrentCD4'], errors='coerce'
            )

            df_ART19['Whostage'] = (
                df_ART19['Whostage']
                .astype(str)
                .str.strip()
                .str.upper()
            )

            df_ART19['ART19'] = (
                df_ART19['Whostage'].isin(
                    ['3', '4', 'III', 'IV', 'STAGE III', 'STAGE IV']
                ) |
                (df_ART19['CurrentCD4'] < 200) |
                (df_ART19['Current_CD4_LFA_Result'] == 'LessThan200')
            ).astype(int)

            ART19Summary = df_ART19.pivot_table(
                index='Sex',
                columns='Age Band',
                values='ART19',
                aggfunc='sum',
                fill_value=0,
                observed=False
            )

            ART19Summary = standardize_ageband_pivot(ART19Summary)
            ART19Summary['Total'] = ART19Summary.sum(axis=1)

        except Exception as e:
            logging.exception("Error Processing ART 19 Summary")
            return jsonify({'error': str(e)}), 500
        
        
        try:
            # ART20 population = ART19 clients
            df_ART20 = df_ART19[df_ART19['ART19'] == 1].copy()

            df_ART20['Serology_for_CrAg_Result'] = (
                df_ART20['Serology_for_CrAg_Result']
                .astype(str)
                .str.strip()
            )
            
            df_ART20['CSF_for_CrAg_Result'] = (
                df_ART20['CSF_for_CrAg_Result']
                .astype(str)
                .str.strip()
            )
            
            # Create a new column 'CrAgNegative' that is 1 if either CrAg result is 'Negative', otherwise 0
            df_ART20['CrAgNegative'] = (
                (df_ART20['Serology_for_CrAg_Result'] == 'Negative') | (df_ART20['CSF_for_CrAg_Result'] == 'Negative')
            ).astype(int)

            ART20aSummary = df_ART20.pivot_table(
                index='Sex',
                columns='Age Band',
                values='CrAgNegative',
                aggfunc='sum',
                fill_value=0,
                observed=False
            )

            ART20aSummary = standardize_ageband_pivot(ART20aSummary)
            ART20aSummary['Total'] = ART20aSummary.sum(axis=1)
            
            
            df_ART20['CrAgPositive'] = (
                (df_ART20['Serology_for_CrAg_Result'] == 'Positive') | (df_ART20['CSF_for_CrAg_Result'] == 'Positive')  
            ).astype(int)

            ART20bSummary = df_ART20.pivot_table(
                index='Sex',
                columns='Age Band',
                values='CrAgPositive',
                aggfunc='sum',
                fill_value=0,
                observed=False
            )

            ART20bSummary = standardize_ageband_pivot(ART20bSummary)
            ART20bSummary['Total'] = ART20bSummary.sum(axis=1)
        
        except Exception as e:
            logging.exception("Error Processing ART 20 Summary")
            return jsonify({'error': str(e)}), 500
        
        try:
            #ART15aSUMMARY
            # Ensure dates are datetime
            df['ARTStartDate'] = pd.to_datetime(df['ARTStartDate'])
            df['First_TPT_Pickupdate'] = pd.to_datetime(df['First_TPT_Pickupdate'])

            # Filter only active clients
            df_everTPT = df#[df['CurrentARTStatus'] == "Active"].copy()

            # Mark as everTPT if started TPT and ARTStartDate is more than 12 months ago
            #df_everTPT['everTPT'] = df_everTPT.apply(
                #lambda row: 1 if (pd.notna(row['First_TPT_Pickupdate']) or pd.notna(row['Current_TPT_Received'])) and (pd.notna(row['ARTStartDate']) and row['ARTStartDate'].to_period('M') > last_year) else 0,
                #axis=1
            #)
            
            # Mark as everTPT if TPT started in month in review and ART was initiated in the last 12 months
            df_everTPT['everTPT'] = df_everTPT.apply(
                lambda row: 1 if ((pd.notna(row['First_TPT_Pickupdate']) and row['First_TPT_Pickupdate'].to_period('M') == Period)) and (pd.notna(row['ARTStartDate']) and row['ARTStartDate'].to_period('M') > last_year) else 0,
                axis=1
            )
            
            #df_everTPT.to_excel('df_everTPT.xlsx')

            # Create pivot table
            ART15aSummary = df_everTPT.pivot_table(
                index='Sex',
                columns='Age Band',
                values='everTPT',
                aggfunc='sum',
                fill_value=0,
                observed=False
            )

            # Standardize index and add totals
            ART15aSummary = ART15aSummary.reindex(['M', 'F'])
            ART15aSummary.rename(index={'M': 'Male', 'F': 'Female'}, inplace=True)
            ART15aSummary['Total'] = ART15aSummary.sum(axis=1)

            # Display summary
            #ART15aSummary
            
        except Exception as e:
            logging.exception("Error Processing ART 15a Summary")
            return jsonify({'error': str(e)}), 500
        
        
        try:
            #ART15bSUMMARY
            # Filter only active clients
            df_everTPT = df#[df['CurrentARTStatus'] == "Active"].copy()

            # Mark as everTPT if started TPT and ARTStartDate is more than 12 months ago
            #df_everTPT['everTPT'] = df_everTPT.apply(
                #lambda row: 1 if (pd.notna(row['First_TPT_Pickupdate']) or pd.notna(row['Current_TPT_Received'])) and (pd.notna(row['ARTStartDate']) and row['ARTStartDate'].to_period('M') <= last_year) else 0,
                #axis=1
            #)
            
            # Mark as everTPT if started TPT and ARTStartDate is more than 12 months ago
            df_everTPT['everTPT'] = df_everTPT.apply(
                lambda row: 1 if ((pd.notna(row['First_TPT_Pickupdate'])) and row['First_TPT_Pickupdate'].to_period('M') == Period) and (pd.notna(row['ARTStartDate']) and row['ARTStartDate'].to_period('M') <= last_year) else 0,
                axis=1
            )
            #df_everTPT.to_excel('df_everTPT.xlsx')

            # Create pivot table
            ART15bSummary = df_everTPT.pivot_table(
                index='Sex',
                columns='Age Band',
                values='everTPT',
                aggfunc='sum',
                fill_value=0,
                observed=False
            )

            # Standardize index and add totals
            ART15bSummary = ART15bSummary.reindex(['M', 'F'])
            ART15bSummary.rename(index={'M': 'Male', 'F': 'Female'}, inplace=True)
            ART15bSummary['Total'] = ART15bSummary.sum(axis=1)

            # Display summary
            #ART15bSummary
            
        except Exception as e:
            logging.exception("Error Processing ART 10 Summary")
            return jsonify({'error': str(e)}), 500
        
        try:
            #ART 16a Summary
            # Filter only active clients
            df_compTPT = df[df['CurrentARTStatus'] == "Active"].copy()

            # Mark clients who have completed at least 6 months on TPT
            df_compTPT['compTPT'] = df_compTPT.apply(
                lambda row: 1 if (
                    (pd.notna(row['First_TPT_Pickupdate'])) and
                    (pd.notna(row['First_TPT_Pickupdate']) and row['First_TPT_Pickupdate'].to_period('M') == last_6mths) and
                    (pd.notna(row['ARTStartDate']) and row['ARTStartDate'].to_period('M') > last_year)
                ) else 0,
                axis=1
            )
            #df_compTPT.to_excel('df_compTPT.xlsx')

            # Create pivot table: ART16a (Active + completed 6 months on TPT)
            ART16aSummary = df_compTPT.pivot_table(
                index='Sex',
                columns='Age Band',
                values='compTPT',
                aggfunc='sum',
                fill_value=0,
                observed=False
            )

            # Standardize index and add row totals
            ART16aSummary = ART16aSummary.reindex(['M', 'F'])
            ART16aSummary.rename(index={'M': 'Male', 'F': 'Female'}, inplace=True)
            ART16aSummary['Total'] = ART16aSummary.sum(axis=1)

            # Optional: Add a total row
            #ART16aSummary.loc['Total'] = ART16aSummary.sum(numeric_only=True)

            #ART16aSummary
            
        except Exception as e:
            logging.exception("Error Processing ART 10 Summary")
            return jsonify({'error': str(e)}), 500
        
        try:            
            #ART 16b Summary
            # Filter only active clients
            df_compTPT = df[df['CurrentARTStatus'] == "Active"].copy()

            # Mark clients who have completed at least 6 months on TPT
            df_compTPT['compTPT'] = df_compTPT.apply(
                lambda row: 1 if (
                    (pd.notna(row['First_TPT_Pickupdate'])) and
                    (pd.notna(row['First_TPT_Pickupdate']) and row['First_TPT_Pickupdate'].to_period('M') == last_6mths) and
                    (pd.notna(row['ARTStartDate']) and row['ARTStartDate'].to_period('M') <= last_year)
                ) else 0,
                axis=1
            )

            # Create pivot table: ART16a (Active + completed 6 months on TPT)
            ART16bSummary = df_compTPT.pivot_table(
                index='Sex',
                columns='Age Band',
                values='compTPT',
                aggfunc='sum',
                fill_value=0,
                observed=False
            )

            # Standardize index and add row totals
            ART16bSummary = ART16bSummary.reindex(['M', 'F'])
            ART16bSummary.rename(index={'M': 'Male', 'F': 'Female'}, inplace=True)
            ART16bSummary['Total'] = ART16bSummary.sum(axis=1)

            # Optional: Add a total row
            #ART16aSummary.loc['Total'] = ART16aSummary.sum(numeric_only=True)
            #ART16bSummary
            
        except Exception as e:
            logging.exception("Error Processing ART 10 Summary")
            return jsonify({'error': str(e)}), 500
        
        
        try:
            # Define the dataframes
            dataframes = {
                "ART2Summary": ART2Summary,
                "ART3Summary": ART3Summary,
                "ART3aSummary": ART3aSummary,
                "ART3bSummary": ART3bSummary,
                "ART3cSummary": ART3cSummary,
                "ART5summary": ART5summary,
                "VLRoutine": VLRoutine,
                "VLTargeted": VLTargeted,
                "VLRoutine_Sup": VLRoutine_Sup,
                "VLTargeted_Sup": VLTargeted_Sup,
                "ART8Summary": ART8Summary,
                "ART9Summary": ART9Summary,
                "ART10aSummary": ART10aSummary,
                "ART10bSummary": ART10bSummary,
                "ART15aSummary": ART15aSummary,
                "ART15bSummary": ART15bSummary,
                "ART16aSummary": ART16aSummary,
                "ART16bSummary": ART16bSummary,
                "ART19Summary": ART19Summary,
                "ART20aSummary": ART20aSummary,
                "ART20bSummary": ART20bSummary
            }

            # Define a title mapping for each sheet
            sheet_titles = {
                "ART2Summary": "ART 2: Number of people living with HIV newly started on ART during the month (excludes ART transfer-in)",
                "ART3Summary": "ART 3: Number of people living with HIV who are currently receiving ART during the month (All regimens)",
                "ART3aSummary": "ART 3 Regimen Lines",
                "ART3bSummary": "ART 3 Multi-Month Dispensing",
                "ART3cSummary": "ART 3 DSD Model",
                "ART5summary": "ART 5: Number of PLHIV on ART who had no clinical contact since their last expected contact.",
                "VLRoutine": "ART 6: Number of PLHIV on ART for at least 6 months with a VL test result during the month - Routine",
                "VLTargeted": "ART 6: Number of PLHIV on ART for at least 6 months with a VL test result during the month - Targeted",
                "VLRoutine_Sup": "ART 7: Number of PLHIV on ART (for at least 6 months) who have virologic suppression (<1000 copies/ml) during the month - Routine",
                "VLTargeted_Sup": "ART 7: Number of PLHIV on ART (for at least 6 months) who have virologic suppression (<1000 copies/ml) during the month - Targeted",
                "ART8Summary": "ART 8: Number of PLHIV who RESTARTED ART during the month",
                "ART9Summary": "ART 9: Number of PLHIV who were TRANSFERRED IN during the month ",
                "ART10aSummary": "ART 10: Number of PLHIV on ART (Including PMTCT) who were Clinically Screened for TB in HIV Treatment Settings - Newly initiated on ART",
                "ART10bSummary": "ART 10: Number of PLHIV on ART (Including PMTCT) who were Clinically Screened for TB in HIV Treatment Settings - Previously on ART",
                "ART15aSummary": "ART15: Number of PLHIV on ART who initiated TB preventive treatment (TPT) - Initiated ART in the last 12 months",
                "ART15bSummary": "ART15: Number of PLHIV on ART who initiated TB preventive treatment (TPT) - On ART greater than 12 months",
                "ART16aSummary": "ART16: Number of PLHIV on ART who completed a course of TB preventive treatment among those who initiated TPT - Initiated ART in the last 12 months (TPT > 6 months assumed)",
                "ART16bSummary": "ART16: Number of PLHIV on ART who completed a course of TB preventive treatment among those who initiated TPT - On ART greater than 12 months (TPT > 6 months assumed)",
                "ART19Summary": "ART 19: Number of newly enrolled PLHIV with WHO clinical stages 3 and 4 and/or CD4 <200 cells/mm3 (Advanced HIV Disease)",
                "ART20aSummary": "ART 20a: Number of newly enrolled PLHIV presenting with WHO clinical stages 3 and 4 and/or CD4<200c/mm3 screened for serum Cryptococcal Antigen (Serum CrAg) before ART initiation - CrAg Negative",
                "ART20bSummary": "ART 20b: Number of newly enrolled PLHIV presenting with WHO clinical stages 3 and 4 and/or CD4<200c/mm3 screened for serum Cryptococcal Antigen (Serum CrAg) before ART initiation - CrAg Positive"
            }

            # Create a new workbook and add a worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "ART MSF"

            def append_df_with_title(ws, title, df, start_row):
                
                # Flatten MultiIndex columns if present
                if isinstance(df.columns, pd.MultiIndex):
                    df.columns = [' | '.join(map(str, col)).strip() for col in df.columns.values]
                
                # Merge the title across all columns
                total_cols = len(df.columns) + 2  # +1 for index, +1 for data columns
                ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=total_cols)
                title_cell = ws.cell(row=start_row, column=1)
                title_cell.value = title
                title_cell.font = Font(bold=True, size=10)
                title_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                start_row += 1
                
                # Add header row (including the index column)
                header = ['Category'] + list(df.columns)  # Add 'Index' as the first column header
                for col_num, value in enumerate(header):
                    cell = ws.cell(row=start_row, column=col_num + 1)
                    cell.value = value
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = PatternFill(start_color="D7E4BC", end_color="D7E4BC", fill_type="solid")
                
                # Add dataframe content (including the index)
                for row_num, (index, row) in enumerate(df.iterrows(), start=start_row + 1):
                    ws.cell(row=row_num, column=1).value = index  # Add index as the first column
                    for col_num, value in enumerate(row):
                        cell = ws.cell(row=row_num, column=col_num + 2)  # Data starts from column 2
                        cell.value = value
                
                # Alternating row color
                for row_num in range(start_row + 1, start_row + len(df) + 1):
                    if row_num % 2 == 0:
                        for col_num in range(1, total_cols + 1):
                            cell = ws.cell(row=row_num, column=col_num)
                            cell.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
                
                return start_row + len(df) + 2


            # Add general title at the top of the sheet
            general_title = f"{facilities_text} ART Monthly Summary Form As At {formatted_period}"
            ws.merge_cells('A1:Q1')  # Merge cells from A1 to Q1 (adjust columns as needed)
            # Set the general title in the merged cell
            general_title_cell = ws.cell(row=1, column=1)
            general_title_cell.value = general_title
            general_title_cell.font = Font(bold=True, size=20)
            general_title_cell.alignment = Alignment(horizontal='center', vertical='center')
            general_title_cell.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
            ws.freeze_panes = 'A2'

            # Process and append each sheet from dataframes
            start_row = 2
            for sheet_name, df in dataframes.items():
                title = sheet_titles.get(sheet_name, sheet_name)
                start_row = append_df_with_title(ws, f">>> {title}", df, start_row)

            # Set column widths dynamically
            for col in range(1, ws.max_column + 1):
                max_length = 0
                column = chr(64 + col)  # Get the column letter (A, B, C, etc.)
                
                # Special case: set fixed width for the first 5 columns (adjust as necessary)
                if col <= 5:
                    adjusted_width = 10  # Set a fixed width for the first 5 columns

                if col <= 1:
                    adjusted_width = 45  # Set a fixed width for the first 5 columns
                #elif col <= 1:
                    #adjusted_width = 45  # Set a fixed width for the first column
                else:
                    for row in ws.iter_rows(min_col=col, max_col=col):
                        for cell in row:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                    adjusted_width = max_length + 2  # Add some padding
                
                # Apply the width to the column
                ws.column_dimensions[column].width = adjusted_width

            # Remove gridlines (this is done by hiding the gridlines in the sheet view)
            ws.sheet_view.showGridLines = False

            old_msf_folder = os.path.join(
                app.config["DOWNLOAD_FOLDER"],
                "old_msf"
            )
            os.makedirs(old_msf_folder, exist_ok=True)

            filename = f"ART MSF SUMMARY AS AT {formatted_period}.xlsx"

            output_path = os.path.join(old_msf_folder, filename)

            wb.save(output_path)

            return jsonify({
                "message": "Old MSF generated successfully.",
                "download_url": url_for(
                    "download_file",
                    filename=f"old_msf/{filename}"
                )
            }), 200

        except Exception as e:
            logging.exception("Error Formating and Exporting data")
            return jsonify({'error': str(e)}), 500

    except Exception as e:
        return jsonify({"message": f"Error processing Excel file: {str(e)}"}), 500
    

# ==============================================================================
# 1. HELPER FUNCTIONS FOR BREAKING DOWN THE WORK
# ==============================================================================

def load_all_files(file1, file2, file3):
    """Loads files and handles the basic merges just like the original code."""
    df = load_file(file1, columns_to_read=columns_to_read)

    # Merge baseline data if present
    if file2:
        df_baseline = load_file(file2, columns_to_read=b_columns_to_read)
        if 'uuid' in df.columns and 'uuid' in df_baseline.columns and 'CurrentARTStatus' in df_baseline.columns:
            df = df.merge(
                df_baseline[['uuid', 'CurrentARTStatus']],
                on='uuid', how='left', suffixes=('', '_baseline')
            )
            df['ARTStatus_PreviousQuarter'] = df['CurrentARTStatus_baseline']
            
    # Merge EMR data if present
    if file3:
        dfbaselineRadet = load_file(file3, columns_to_read=r_columns_to_read)
        df = process_emr_data(df, dfbaselineRadet, emr_df)
        
    return df


def calculate_age_bands(df, end_date):
    """Calculates age and adds age bands using your original vectorized logic."""
    if end_date is None:
        today = pd.Timestamp.today().normalize()
    else:
        today = pd.to_datetime(end_date) + pd.offsets.MonthEnd(0)

    dob = df['DOB']
    df['Age'] = (today.year - dob.dt.year 
                 - ((dob.dt.month > today.month) | 
                    ((dob.dt.month == today.month) & (dob.dt.day > today.day))).astype(int))
    
    # Run the common age-band helper
    df = msf_common.add_agebands(df)
    return df


def get_age_summary(df_source, value_column):
    """Your exact template-matching pivot table function."""
    pt = df_source.pivot_table(
        index="Sex",
        columns="Age Band New",
        values=value_column,
        aggfunc="sum",
        fill_value=0,
        observed=False
    )
    return msf_common.standardize_pivot(pt, msf_common.NEW_AGE_BANDS)


def write_grid_to_excel(ws, summary_table, start_row):
    """
    Fills out a standard 2-row grid (Male and Female) across columns C to J.
    Instead of hardcoding 16 cells for every section, it uses a simple loop!
    """
    columns = ["C", "D", "E", "F", "G", "H", "I", "J"]
    age_bands = ["<1", "1-4", "5-9", "10-14", "15-19", "20-24", "25-49", "50+"]
    
    for col_letter, age in zip(columns, age_bands):
        # Write Male row
        try:
            ws[f"{col_letter}{start_row}"] = int(summary_table.loc["Male", age])
        except Exception:
            ws[f"{col_letter}{start_row}"] = 0
            
        # Write Female row (1 row right below Male)
        try:
            ws[f"{col_letter}{start_row + 1}"] = int(summary_table.loc["Female", age])
        except Exception:
            ws[f"{col_letter}{start_row + 1}"] = 0


# ==============================================================================
# 2. THE CLEAN FLASK ROUTE
# ==============================================================================

@app.route("/fetch_newmsf", methods=["POST"])
def fetch_new_msf():
    try:
        
        file1 = request.files.get("file1")
        file2 = request.files.get("file2")
        file3 = request.files.get("file3")
        
        # --- Phase 1: File Loading & Cleaning ---
        df = load_all_files(file1, file2, file3)
        
        for col in DATE_COLUMNS:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors='coerce', dayfirst=True)
                
        for col in NUMERIC_COLUMNS:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')

        # --- Phase 2: Date Parsing & Metadata ---
        end_date = request.form.get("endDate")
        global formatted_period
        if end_date:
            end_date_dt = pd.to_datetime(end_date)
            formatted_period = end_date_dt.to_period('M').strftime('%B %Y')
            Period = end_date_dt.to_period('M')

        unique_facilities = df['FacilityName'].unique()
        facilities_text = ', '.join(unique_facilities)
        
        # --- Phase 3: Feature Engineering ---
        df = calculate_age_bands(df, end_date)

        # Map simple integer flags
        df["IsPregnant"] = (df["CurrentPregnancyStatus"] == "Pregnant").astype(int)
        df["IsBreastfeeding"] = (df["CurrentPregnancyStatus"] == "Breastfeeding").astype(int)
        df["OnTB"] = (df["Current_TB_Status"] == "On treatment for disease").astype(int)

        df["IsFirstLine"] = df["CurrentRegimenLine"].isin(["Adult 1st line ARV regimen", "Child 1st line ARV regimen"]).astype(int)
        df["IsSecondLine"] = df["CurrentRegimenLine"].isin(["Adult 2nd line ARV regimen", "Child 2nd line ARV regimen"]).astype(int)
        df["IsThirdLine"] = df["CurrentRegimenLine"].isin(["Adult 3rd Line ARV Regimens", "Child 3rd line ARV regimen"]).astype(int)

        # --- Phase 4: Slicing the Cohorts ---
        df_active = df[df["CurrentARTStatus"] == "Active"].copy()
        
        df_new = df[
            (df["ARTStartDate"].dt.to_period("M") == Period) &
            (df["CurrentARTStatus"] == "Active")
        ].copy()
        
        
        # ===========================
        # Losses
        # ===========================
        df['Pharmacy_LastPickupdate2'] = pd.to_datetime(
            df['Pharmacy_LastPickupdate'], errors='coerce'
        ).fillna(pd.to_datetime('1900-01-01'))

        df['DaysOfARVRefill2'] = df['DaysOfARVRefill'].apply(
            lambda x: 0 if pd.notna(x) and x > 180 else x
        )

        df['NextAppt'] = (
            df['Pharmacy_LastPickupdate2'] +
            pd.to_timedelta(df['DaysOfARVRefill2'].fillna(0), unit='D')
        )

        df['IITDate2'] = df['NextAppt'] + pd.Timedelta(days=29)

        df['Losses date'] = df['Outcomes_Date'].fillna(df['IITDate2'])
        df['Losses date'] = pd.to_datetime(df['Losses date'], errors='coerce')

        df_Losses = df[
            df['CurrentARTStatus'].isin([
                "Death",
                "Transferred out",
                "LTFU",
                "Discontinued Care"
            ]) &
            (df['Losses date'].dt.to_period('M') == Period)
        ].copy()

        df_Losses["Stopped"] = (df_Losses["CurrentARTStatus"] == "Discontinued Care").astype(int)
        df_Losses["LTFU"] = (df_Losses["CurrentARTStatus"] == "LTFU").astype(int)
        df_Losses["Dead"] = (df_Losses["CurrentARTStatus"] == "Death").astype(int)
        df_Losses["Transferred"] = (df_Losses["CurrentARTStatus"] == "Transferred out").astype(int)
        
        df['ARTStartDate'] = pd.to_datetime(df['ARTStartDate'], errors='coerce')

        df_Restart = df[
            (df['CurrentARTStatus'] == "Active") &
            (df['ARTStartDate'].dt.to_period('M') != Period)
        ].copy()

        df_Restart['Restart'] = np.where(
            (df_Restart['ARTStatus_PreviousQuarter'].notna()) &
            (df_Restart['ARTStatus_PreviousQuarter'] != "Active") &
            (df_Restart['Date_Transfered_In'].dt.to_period('M') != Period),
            1,
            0
        )

        # --- Phase 5: Generating Summaries & Excel Writing ---
        template_path = os.path.join(app.root_path, "templates", "New_MSF.xlsx")
        wb = load_workbook(template_path)
        ws = wb.active

        # Header block
        ws.merge_cells("A1:M1")
        ws["A1"] = f"{facilities_text} - {formatted_period} (Generated: {datetime.now():%d-%b-%Y %H:%M})"
        ws["A1"].font = Font(bold=True, size=14, color="FFFFFFFF")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        # Write Age/Sex Grid Summaries using our simple row-helper
        df_new["ART1"] = 1
        write_grid_to_excel(ws, get_age_summary(df_new, "ART1"), start_row=4)
        
        df_active["ART2"] = 1
        write_grid_to_excel(ws, get_age_summary(df_active[df_active["IsFirstLine"] == 1], "ART2"), start_row=13)
        write_grid_to_excel(ws, get_age_summary(df_active[df_active["IsSecondLine"] == 1], "ART2"), start_row=17)
        write_grid_to_excel(ws, get_age_summary(df_active[df_active["IsThirdLine"] == 1], "ART2"), start_row=21)

        # Viral Load Queries (ART3 & ART4)
        last_year = (Period.to_timestamp() - pd.DateOffset(months=12)).to_period('M')
        last_6mths = (Period.to_timestamp() - pd.DateOffset(months=6)).to_period('M')
        
        df_VL = df_active[
            (df_active['DateResultReceivedFacility'].dt.to_period('M') > last_year) &
            (df_active['ARTStartDate'].dt.to_period('M') <= last_6mths)
        ].copy()
        df_VL["ART3"] = 1
        write_grid_to_excel(ws, get_age_summary(df_VL, "ART3"), start_row=31)

        df_VL_Sup = df_VL[df_VL['CurrentViralLoad'] < 1000].copy()
        df_VL_Sup["ART4"] = 1
        write_grid_to_excel(ws, get_age_summary(df_VL_Sup, "ART4"), start_row=39)

        # --- Phase 6: KPI Population Counts ---
        # Simple local helper function to count specific categories to save space
        def count_kp(dataframe, kp_string):
            return int((dataframe["KPType"] == kp_string).sum())

        # ART1 KPs
        ws["C8"], ws["D8"], ws["E8"] = count_kp(df_new, "Male who has sex with men"), count_kp(df_new, "FSW"), count_kp(df_new, "PWID")
        ws["F8"], ws["G8"] = count_kp(df_new, "Transgender"), count_kp(df_new, "In prison")

        # ART2 KPs & MMD Refills
        ws["K10"] = df_active.shape[0]
        ws["C25"], ws["D25"], ws["E25"] = count_kp(df_active, "Male who has sex with men"), count_kp(df_active, "FSW"), count_kp(df_active, "PWID")
        ws["F25"], ws["G25"] = count_kp(df_active, "Transgender"), count_kp(df_active, "In prison")

        ws["C27"] = (df_active["DaysOfARVRefill"] < 90).sum()
        ws["E27"] = ((df_active["DaysOfARVRefill"] >= 90) & (df_active["DaysOfARVRefill"] < 180)).sum()
        ws["H27"] = (df_active["DaysOfARVRefill"] >= 180).sum()

        # ART3 & ART4 KPs
        ws["C35"], ws["D35"], ws["E35"] = count_kp(df_VL, "Male who has sex with men"), count_kp(df_VL, "FSW"), count_kp(df_VL, "PWID")
        ws["F35"], ws["G35"] = count_kp(df_VL, "Transgender"), count_kp(df_VL, "In prison")

        ws["C43"], ws["D43"], ws["E43"] = count_kp(df_VL_Sup, "Male who has sex with men"), count_kp(df_VL_Sup, "FSW"), count_kp(df_VL_Sup, "PWID")
        ws["F43"], ws["G43"] = count_kp(df_VL_Sup, "Transgender"), count_kp(df_VL_Sup, "In prison")
        
        #ART 5
        ws["C46"] = (df_Losses["CurrentARTStatus"] == "Discontinued Care").sum()
        ws["F46"] = (df_Losses["CurrentARTStatus"] == "LTFU").sum()
        ws["I46"] = (df_Losses["CurrentARTStatus"] == "Death").sum()
        ws["K49"] = (df_Losses["CurrentARTStatus"] == "Transferred out").sum()
        
        #Restart and Transfer In
        ws["K48"] = df_Restart["Restart"].sum()
        ws["K50"] = (
            (df["CurrentARTStatus"] == "Active") &
            (df["Date_Transfered_In"].dt.to_period("M") == Period)
        ).sum()

        # --- Phase 7: Save File Output ---
        safe_facility = facilities_text.replace("/", "-")
        new_msf_folder = os.path.join(app.config["DOWNLOAD_FOLDER"], "new_msf")
        os.makedirs(new_msf_folder, exist_ok=True)

        filename = f"{safe_facility}_NEW_MSF_{formatted_period.replace(' ', '_')}.xlsx"
        output_path = os.path.join(new_msf_folder, filename)
        wb.save(output_path)

        return jsonify({
            "message": "New MSF generated successfully.",
            "download_url": url_for("download_file", filename=f"new_msf/{filename}")
        }), 200

    except Exception as e:
        logging.exception("Error Processing Data")
        return jsonify({'error': str(e)}), 500

@app.route("/download/<path:filename>")
def download_file(filename):
    download_root = app.config["DOWNLOAD_FOLDER"]

    file_path = safe_join(download_root, filename)

    if file_path is None or not os.path.isfile(file_path):
        abort(404, description="File not found")

    return send_from_directory(
        download_root,
        filename,
        as_attachment=True
    )

if __name__ == '__main__':
    app.run(debug=True)
